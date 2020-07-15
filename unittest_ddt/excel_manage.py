# -*- coding: utf-8 -*-
# @Time : 2020/6/16 10:22 下午
# @Author : zhl
# @File : excel_manage.py
# @Software: PyCharm


from openpyxl import load_workbook
import  os
#得到路径
file_path= os.path.join(os.path.dirname(os.path.abspath(__file__)),'unittest.xlsx')

print(file_path)
#生成workbook对象（wb）
wb =load_workbook(file_path)
#根据
sh =wb['login']
print(sh)
#
cel = sh.cell(2,2)
print(cel.value)

print(sh.max_row)
print(sh.max_column)

sh=wb['login']
titles=[]

print(list(sh.rows))#每一行是个元组，元组里放到是每一行到单元格

for item in list(sh.rows):
    #[(<Cell 'login'.A1>, <Cell 'login'.B1>, <Cell 'login'.C1>), (<Cell 'login'.A2>, <Cell 'login'.B2>, <Cell 'login'.C2>), (<Cell 'login'.A3>, <Cell 'login'.B3>, <Cell 'login'.C3>)]
    for cel in item:#获取每一行单元格数据
        print(cel.value)